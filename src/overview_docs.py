import json
import os
from calendar import Calendar

import openpyxl
from openpyxl.styles import Alignment, Font

from common import Querier, get_logger
from config_share import *

__version__ = '1.0.1'

logger = get_logger(PATH_LOG, __name__)


def generate_overview(year: int = date.today().year, month: int = date.today().month) -> None:
    """
    Generate the month overview Excel for specific year and month deliveries.

    :param year: The desired year for the overview.
    :type year: int
    :param month: The desired month for the overview.
    :type month: int
    """
    querier: Querier = Querier()
    # overview_name: path to the result overview (es. c:/source/wheeck/res/2024_01.xlsx)
    overview_name = f'{PATH_RES}/{year}_{month:0>2}.xlsx'
    # scheme_name: path to the overview template (es. c:/source/wheeck/scheme/overview.xlsx)
    scheme_name = f'{PATH_SCHEME}/overview.xlsx'

    if querier.run(QUERY_GET_OVERVIEW, (year, month)).rows:
        # wb: raw XLSX doc
        wb = openpyxl.load_workbook(overview_name if os.path.isfile(overview_name) else scheme_name)

        # ws: raw XLSX doc sheet
        ws = wb['consegne']
        for row_num, row in enumerate(querier, start=2):
            for col_num, col in enumerate(row, start=1):
                ws.cell(row=row_num, column=col_num).value = col
                ws.cell(row=row_num, column=col_num).font = Font(name=OVERVIEW_DEFAULT_FONT)
                ws.cell(row=row_num, column=col_num).number_format = OVERVIEW_FORMATS[type(col)]

        # updating date column in other workbook sheets
        for ws in [wb['cifre'], wb['litri'], wb['cifre manuale'], wb['litri manuale']]:
            ws.cell(row=1, column=1).value = year
            ws.cell(row=1, column=1).font = Font(name=OVERVIEW_DEFAULT_FONT)
            ws.cell(row=1, column=1).number_format = OVERVIEW_FORMATS[str]

            for row_num, row in enumerate([day for day in Calendar().itermonthdates(year, month) if day.month == month], start=3):
                ws.cell(row=row_num, column=1).value = row
                ws.cell(row=row_num, column=1).font = Font(name=OVERVIEW_DEFAULT_FONT)
                ws.cell(row=row_num, column=1).number_format = OVERVIEW_FORMATS[date]

        logger.info(f'saving overview for date {year}-{month:0>2}... [{overview_name.split('/')[-1]}]')
        wb.save(overview_name)
    else:
        logger.info(f'no delivery found in date {year}-{month:0>2}... skipping overview!')
    del querier


def update_overviews() -> None:
    """
    Update or generate the overviews of deliveries recorded today.
    """
    querier: Querier = Querier()
    if querier.run(QUERY_GET_LAST_OVERVIEWS).rows:
        for row in querier:
            generate_overview(row.year, row.month)
    del querier


def generate_summary(year: int = date.today().year) -> None:
    """
    Generate the year trip summary Excel for specific year deliveries.

    :param year: The desired year for the summary.
    :type year: int
    """
    querier: Querier = Querier()
    # summary_name: path to the result overview (es. c:/source/wheeck/res/2024_TRIPS.xlsx)
    summary_name = f'{PATH_RES}/{year}_TRIPS.xlsx'
    # scheme_name: path to the overview template (es. c:/source/wheeck/scheme/summary.xlsx)
    scheme_name = f'{PATH_SCHEME}/summary.xlsx'

    with open(PATH_CFG_PRJ) as jin:
        members = json.load(jin)['members']
    drivers = [e['driver'] for e in members if not e['vehicle']]
    vehicles = sorted([e['vehicle'] for e in members if e['vehicle']])

    if querier.run(QUERY_GET_SUMMARY % ', '.join('?' * len(drivers)), (year, *drivers, *vehicles)).rows:
        # wb: raw XLSX doc
        wb = openpyxl.load_workbook(scheme_name)

        # ws: raw XLSX doc sheet
        ws = wb['viaggi']
        for row_num, row in enumerate(querier, start=3):
            for col_num, col in enumerate(row, start=1):
                ws.cell(row=row_num, column=col_num).value = col
                ws.cell(row=row_num, column=col_num).font = Font(name=OVERVIEW_DEFAULT_FONT)
                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center')
                ws.cell(row=row_num, column=col_num).number_format = OVERVIEW_FORMATS[type(col)]

        logger.info(f'saving summary for year {year}... [{summary_name.split('/')[-1]}]')
        wb.save(summary_name)
    else:
        logger.info(f'no delivery found in year {year}... skipping summary!')
    del querier


def update_summaries() -> None:
    """
    Update or generate the summaries of deliveries recorded today.
    """
    querier: Querier = Querier()
    if querier.run(QUERY_GET_LAST_OVERVIEWS).rows:
        for year in set([row[0] for row in querier.fetch(Querier.FETCH_ALL)]):
            generate_summary(year)
    del querier
