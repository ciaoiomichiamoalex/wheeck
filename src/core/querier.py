import sqlite3
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, Self

import openpyxl
import pyodbc
from openpyxl.styles import Font

from .common import decode_json


class Querier:
    """
    The Querier object allows for run queries on the database and fetch the extracted data.
    """
    _cache_config: dict[str, dict] = {}

    FETCH_VAL: int = 10
    FETCH_ONE: int = 20
    FETCH_ALL: int = 30
    FETCH_MANY: int = 40

    _EXCEL_FORMATS: dict = {
        type(None): 'General',
        str: '@',
        int: '#,##0',
        float: '#,##0.00',
        Decimal: '#,##0.00',
        date: 'dd/mm/yyyy',
        time: 'h:mm:ss;@',
        datetime: 'dd/mm/yyyy h:mm:ss;@'
    }

    def __init__(self,
                 cfg_in: str | Path = None,
                 conn_name: str = 'main',
                 conn_str: dict = None,
                 save_changes: bool = False) -> None:
        """
        Read from a JSON config file the database configuration and start the connection.

        :param cfg_in: The path to the JSON file with the database configurations, defaults to None.
        :type cfg_in: str | Path
        :param conn_name: The database configuration name in the JSON file, defaults to 'main'.
        :type conn_name: str
        :param conn_str: Allow to pass database configuration manually and override conn_name, defaults to None.
        :type conn_str: dict
        :param save_changes: Enable or disable the auto-commit, defaults to False.
        :type save_changes: bool
        :raise IOError: If configuration is not found.
        """
        if not conn_str and cfg_in:
            cfg_in = Path(cfg_in).resolve()

            if cfg_in.is_dir():
                cfg_in = cfg_in / 'querier.json'

            if not cfg_in.is_file():
                raise IOError(f'Querier: no config {cfg_in} found!')

            conn_in = f'{cfg_in}::{conn_name}'
            if not Querier._cache_config or conn_in not in Querier._cache_config:
                Querier._cache_config[conn_in] = decode_json(cfg_in, name=conn_name)
            config = Querier._cache_config.get(conn_in)

            if not config: raise IOError(f'Querier: no config <{conn_name}> found in {cfg_in}!')
        elif conn_str: config = conn_str
        else: raise IOError('Querier: no config found!')

        self._connection: pyodbc.Connection = pyodbc.connect(
            driver=f"{{{config['driver']}}}",
            server=config['server'],
            port=config['port'],
            database=config['database'],
            user=config['user'],
            password=config['password'],
            autocommit=save_changes
        )
        self._cursor: pyodbc.Cursor = self._connection.cursor()
        self.rows: int = 0

    def __del__(self) -> None:
        self._cursor.close()
        self._connection.close()

    def __iter__(self) -> pyodbc.Cursor:
        return self._cursor

    @property
    def cursor(self) -> pyodbc.Cursor:
        """
        Exposes the cursor to make available further calls not wrapped in this class.

        :return: The cursor to the database.
        :rtype: Cursor
        """
        return self._cursor

    def run(self,
            query: str,
            *args) -> Self:
        """
        Execute DDL, DML or DQL query on the database.

        :param query: The query string to be executed.
        :type query: str
        :param args: The parameters list of the query string, as positional arguments or single iterable.
        :type args: Any
        :return: The object itself, so that calls can be chained.
        :rtype: Querier
        """
        self.rows = (
            self._cursor.execute(query, *args).rowcount
            if args and set(args) != {None}
            else self._cursor.execute(query).rowcount
        )
        return self

    def fetch(self,
              genre: int = FETCH_MANY,
              size: int = 200) -> Any:
        """
        Return the last query result based on fetch genre FETCH_VAL, FETCH_ONE, FETCH_ALL or FETCH_MANY.

        :param genre: The class constant fetch genre, defaults to FETCH_MANY.
        :type genre: int
        :param size: The number of rows in case of FETCH_MANY genre, defaults to 200.
        :type size: int
        :return: The query result set.
        :rtype: Any
        """
        match genre:
            case Querier.FETCH_VAL:
                return self._cursor.fetchval()
            case Querier.FETCH_ONE:
                return self._cursor.fetchone()
            case Querier.FETCH_ALL:
                return self._cursor.fetchall()
            case Querier.FETCH_MANY:
                return self._cursor.fetchmany(size)
            case _:
                return None

    def save_changes(self,
                     save: bool = True) -> None:
        """
        Save or revert the changes done on the database.

        :param save: True for commit changes, False for rollback changes, defaults to True.
        :type save: bool
        """
        return self._cursor.commit() if save else self._cursor.rollback()

    def row_header(self) -> list[str] | None:
        """
        Get the column names list of the last query done.

        :return: The column names list or None if there isn't columns.
        :rtype: list[str] | None
        """
        return (
            [column[0] for column in self._cursor.description]
            if self._cursor.description else None
        )

    def save_excel(self,
                   fou: str | Path,
                   sheet_name: str = None,
                   font_face: str = None) -> None:
        """
        Save the last query result into an Excel file.

        :param fou: The path to the result file, including filename.
        :type fou: str | Path
        :param sheet_name: The name of the sheet into the file, defaults to Excel defaults sheet name.
        :type sheet_name: str
        :param font_face: The font name to be used into the file, defaults to Excel defaults font family.
        :type font_face: str
        """
        if self._cursor:
            wb = openpyxl.Workbook()
            ws = wb.active
            if sheet_name: ws.title = sheet_name

            for col_num, cell in enumerate(self.row_header(), start=1):
                ws.cell(row=1, column=col_num).value = cell
                ws.cell(row=1, column=col_num).font = Font(name=font_face, bold=True)
                ws.cell(row=1, column=col_num).number_format = Querier._EXCEL_FORMATS.get(type(cell))

            for row_num, row in enumerate(self._cursor, start=2):
                for col_num, cell in enumerate(row, start=1):
                    ws.cell(row=row_num, column=col_num).value = cell
                    ws.cell(row=row_num, column=col_num).font = Font(name=font_face)
                    ws.cell(row=row_num, column=col_num).number_format = Querier._EXCEL_FORMATS.get(type(cell))

            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = 'A2'
            wb.save(fou)


class LowQuerier(Querier):
    """
    The LowQuerier object allows for run queries on SQLite database and fetch the extracted data.
    """
    def __init__(self,
                 conn_in: str | Path = ':memory:',
                 save_changes: bool = False) -> None:
        """
        Start the connection to the SQLite database.

        :param conn_in: The path to the SQLite database file, defaults to in memory database.
        :type conn_in: str | Path
        :param save_changes: Enables or disables the auto-commit, defaults to True.
        :type save_changes: bool
        """
        self._connection: sqlite3.Connection = (
            sqlite3.connect(database=conn_in, autocommit=save_changes)
            if save_changes else sqlite3.connect(database=conn_in)
        )
        self._connection.row_factory = sqlite3.Row
        self._cursor: sqlite3.Cursor = self._connection.cursor()
        self.rows: int = 0

    def __iter__(self) -> sqlite3.Cursor:
        return super().__iter__()

    @property
    def cursor(self) -> sqlite3.Cursor:
        """
        Exposes the cursor to make available further calls not wrapped in this class.

        :return: The cursor to the database.
        :rtype: Cursor
        """
        return super().cursor

    def run(self,
            query: str,
            *args) -> Self:
        """
        Execute DDL, DML or DQL query on the database.

        :param query: The query string to be executed.
        :type query: str
        :param args: The parameters list of the query string, as positional arguments or single iterable.
        :type args: Any
        :return: The object itself, so that calls can be chained.
        :rtype: Querier
        """
        if len(args) == 1 and isinstance(args[0], (list, tuple, set)):
            args = tuple(args[0])
        return super().run(query, args)

    def fetch(self,
              genre: int = Querier.FETCH_MANY,
              size: int = 200) -> Any:
        """
        Return the last query result based on fetch genre FETCH_VAL, FETCH_ONE, FETCH_ALL or FETCH_MANY.

        :param genre: The class constant fetch genre, defaults to FETCH_MANY.
        :type genre: int
        :param size: The number of rows in case of FETCH_MANY genre, defaults to 200.
        :type size: int
        :return: The query result set.
        :rtype: Any
        """
        res = super().fetch(LowQuerier.FETCH_ONE if genre == LowQuerier.FETCH_VAL else genre, size)
        if not res: return None

        match genre:
            case LowQuerier.FETCH_VAL:
                return res[0]
            case LowQuerier.FETCH_ONE:
                return dict(res)
            case LowQuerier.FETCH_ALL | LowQuerier.FETCH_MANY:
                return [dict(row) for row in res]
            case _:
                return None
