import datetime
import decimal
import json
import logging
from typing import Any, Self

import openpyxl
import pyodbc
from openpyxl.styles import Font

from config_share import PATH_CFG_COMMON

__version__ = '1.0.1'


def decode_json(json_in: str, only_one: bool = True, **kwargs) -> dict | list | None:
    """
    Read a json file and return the objects which verify the condition {key: value}.

    :param json_in: The path to the json file.
    :type json_in: str
    :param only_one: Indicates if returns only the first object or the entire list.
    :type only_one: bool
    :param kwargs: The attribute conditions to be verified.
    :return: A list of matching objects in the file or a single object, or None if no one match.
    :rtype: dict | list | None
    """
    with open(json_in) as jin:
        res = list(json.load(jin))

    res = [obj for obj in res if all(obj.get(key) == value for key, value in kwargs.items())]
    return res[0] if res and only_one else res if res else None


def get_logger(fou: str, name: str = 'main', log_level: int | str = logging.INFO, console: bool = True) -> logging.Logger:
    """
    Initialize a new logger object with custom properties, by creating it if it doesn't already exist.

    :param fou: The path to the log file.
    :type fou: str
    :param name: The name of the logger, defaults to 'main'.
    :type name: str
    :param log_level: The logging level on file, defaults to INFO.
    :type log_level: int | str
    :param console: Enable or disable logging error on console, defaults to True.
    :type console: bool
    :return: The logger object.
    :rtype: Logger
    """
    logger = logging.getLogger(name)
    if not logger.hasHandlers():
        logger.setLevel(logging.DEBUG)

        fou_handler = logging.FileHandler(fou)
        fou_handler.setLevel(log_level)
        fou_handler.setFormatter(logging.Formatter(
            '%(asctime)s [%(name)s] %(levelname)s - %(message)s',
            '%d/%m/%Y %H:%M:%S'
        ))
        logger.addHandler(fou_handler)

        if console:
            console_handler = logging.StreamHandler()
            console_handler.setLevel(logging.ERROR)
            console_handler.setFormatter(logging.Formatter(
                '%(asctime)s [%(name)s] %(levelname)s - %(message)s',
                '%d/%m/%Y %H:%M:%S'
            ))
            logger.addHandler(console_handler)
    return logger


class Querier:
    FETCH_VAL: int = 10
    FETCH_ONE: int = 20
    FETCH_ALL: int = 30
    FETCH_MANY: int = 40

    _EXCEL_DEFAULT_FONT = 'Aptos Narrow'
    _EXCEL_FORMATS = {
        type(None): 'General',
        str: '@',
        int: '#,##0',
        float: '#,##0.00',
        decimal.Decimal: '#,##0.00',
        datetime.date: 'dd/mm/yyyy',
        datetime.time: 'h:mm:ss;@',
        datetime.datetime: 'dd/mm/yyyy h:mm:ss;@'
    }

    def __init__(self, conn_name: str = 'main', save_changes: bool = False) -> None:
        """
        Read from config/common.json the database configuration and start the connection.

        :param conn_name: The database configuration name in common.json, defaults to 'main'.
        :type conn_name: str
        :param save_changes: Enables or disables the auto-commit, defaults to False.
        :type save_changes: bool
        :raise IOError: If configuration name is not found.
        """
        config = decode_json(PATH_CFG_COMMON, name=conn_name, genre='database')
        if not config:
            raise IOError(f'Querier: no config <{conn_name}> found!')

        self._connection = pyodbc.connect(
            driver=f"{{{config['driver']}}}",
            server=config['server'],
            port=config['port'],
            database=config['database'],
            user=config['user'],
            password=config['password'],
            autocommit=save_changes
        )
        self._cursor = self._connection.cursor()
        self.rows: int = 0

    def __del__(self) -> None:
        self._cursor.close()
        self._connection.close()

    def __iter__(self) -> pyodbc.Cursor:
        return self._cursor

    @property
    def cursor(self) -> pyodbc.Cursor:
        """
        Exposes the cursor to make available further calls not wrapped in this class.

        :return: The cursor to the database.
        :rtype: Cursor
        """
        return self._cursor

    def run(self, query: str, args: list | tuple = None) -> Self:
        """
        Execute DQL, DDL or DML query on the database

        :param query: The query string to be executed.
        :type query: str
        :param args: The parameters list of the query string, defaults to None.
        :type args: list | tuple
        :return: The object itself, so that calls can be chained.
        :rtype: Querier
        """
        self.rows = self._cursor.execute(query, args).rowcount if args else self._cursor.execute(query).rowcount
        return self

    def fetch(self, genre: int = FETCH_MANY, size: int = 200) -> Any:
        """
        Return the last query result based on fetch genre FETCH_VAL, FETCH_ONE, FETCH_ALL and FETCH_MANY.

        :param genre: The fetch type, defaults to FETCH_MANY.
        :type genre: int
        :param size: The number of rows in case of FETCH_MANY, defaults to 200.
        :type size: int
        :return: The query result set.
        :rtype: Any
        """
        match genre:
            case x if x == self.FETCH_VAL:
                return self._cursor.fetchval()
            case x if x == self.FETCH_ONE:
                return self._cursor.fetchone()
            case x if x == self.FETCH_ALL:
                return self._cursor.fetchall()
            case x if x == self.FETCH_MANY:
                return self._cursor.fetchmany(size)

    def row_header(self) -> list[str] | None:
        """
        Get the column names list of the last query.

        :return: The column names list or None if there isn't columns.
        :rtype: list[str] | None
        """
        return [column[0] for column in self._cursor.description]

    def save_excel(self, fou: str, sheet_name: str = None, font_face: str = _EXCEL_DEFAULT_FONT) -> None:
        """
        Save the last query result into an Excel file.

        :param fou: Path to the result file.
        :type fou: str
        :param sheet_name: Name of the sheet into the file, defaults to None.
        :type sheet_name: str
        :param font_face: Font name used into the file, defaults to EXCEL_DEFAULT_FONT class constant.
        :type font_face: str
        """
        if self._cursor:
            wb = openpyxl.Workbook()
            ws = wb.active
            if sheet_name:
                ws.title = sheet_name

            for col_num, col in enumerate(self.row_header(), start=1):
                ws.cell(row=1, column=col_num).value = col
                ws.cell(row=1, column=col_num).font = Font(name=font_face, bold=True)
                ws.cell(row=1, column=col_num).number_format = self._EXCEL_FORMATS[type(col)]

            for row_num, row in enumerate(self._cursor, start=2):
                for col_num, col in enumerate(row, start=1):
                    ws.cell(row=row_num, column=col_num).value = col
                    ws.cell(row=row_num, column=col_num).font = Font(name=font_face)
                    ws.cell(row=row_num, column=col_num).number_format = self._EXCEL_FORMATS[type(col)]

            ws.auto_filter.ref = ws.dimensions
            wb.save(fou)
