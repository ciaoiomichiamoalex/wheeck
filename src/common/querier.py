from datetime import date, datetime, time
from decimal import Decimal
from typing import Any, Self

import openpyxl
import pyodbc
from openpyxl.styles import Font

from .common import decode_json
from .constants import PATH_CFG_COMMON


class Querier:
    """
    The Querier object allows for run queries on the database and fetch the extracted data.
    """
    FETCH_VAL: int = 10
    FETCH_ONE: int = 20
    FETCH_ALL: int = 30
    FETCH_MANY: int = 40

    _EXCEL_FORMATS = {
        type(None): 'General',
        str: '@',
        int: '#,##0',
        float: '#,##0.00',
        Decimal: '#,##0.00',
        date: 'dd/mm/yyyy',
        time: 'h:mm:ss;@',
        datetime: 'dd/mm/yyyy h:mm:ss;@'
    }

    def __init__(self, conn_name: str = 'main', save_changes: bool = False, conn_str: dict = None) -> None:
        """
        Read from config/common.json the database configuration and start the connection.

        :param conn_name: The database configuration name in common.json, defaults to 'main'.
        :type conn_name: str
        :param save_changes: Enables or disables the auto-commit, defaults to False.
        :type save_changes: bool
        :param conn_str: Allow to pass database configuration manually, override conn_name.
        :type conn_str: dict
        :raise IOError: If configuration name is not found.
        """
        config = conn_str if conn_str else decode_json(PATH_CFG_COMMON, name=conn_name, genre='database')
        if not config:
            raise IOError(f'Querier: no config <{conn_name}> found!')

        self._connection = pyodbc.connect(
            driver=f"{{{config['driver']}}}",
            server=config['server'],
            port=config['port'],
            database=config['database'],
            user=config['user'],
            password=config['password'],
            autocommit=save_changes
        )
        self._cursor = self._connection.cursor()
        self.rows: int = 0

    def __del__(self) -> None:
        self._cursor.close()
        self._connection.close()

    def __iter__(self) -> pyodbc.Cursor:
        return self._cursor

    @property
    def cursor(self) -> pyodbc.Cursor:
        """
        Exposes the cursor to make available further calls not wrapped in this class.

        :return: The cursor to the database.
        :rtype: Cursor
        """
        return self._cursor

    def run(self, query: str, *args) -> Self:
        """
        Execute DQL, DDL or DML query on the database.

        :param query: The query string to be executed.
        :type query: str
        :param args: The parameters list of the query string, as positional arguments or single iterable.
        :return: The object itself, so that calls can be chained.
        :rtype: Querier
        """
        self.rows = (self._cursor.execute(query, *args).rowcount if args and set(args) != {None}
                     else self._cursor.execute(query).rowcount)
        return self

    def fetch(self, genre: int = FETCH_MANY, size: int = 200) -> Any:
        """
        Return the last query result based on fetch genre FETCH_VAL, FETCH_ONE, FETCH_ALL and FETCH_MANY.

        :param genre: The class constant fetch genre, defaults to FETCH_MANY.
        :type genre: int
        :param size: The number of rows in case of FETCH_MANY genre, defaults to 200.
        :type size: int
        :return: The query result set.
        :rtype: Any
        """
        match genre:
            case self.FETCH_VAL:
                return self._cursor.fetchval()
            case self.FETCH_ONE:
                return self._cursor.fetchone()
            case self.FETCH_ALL:
                return self._cursor.fetchall()
            case self.FETCH_MANY:
                return self._cursor.fetchmany(size)

    def row_header(self) -> list[str] | None:
        """
        Get the column names list of the last query done.

        :return: The column names list or None if there isn't columns.
        :rtype: list[str] | None
        """
        return [column[0] for column in self._cursor.description] if self._cursor.description else None

    def save_excel(self, fou: str, sheet_name: str = None, font_face: str = None) -> None:
        """
        Save the last query result into an Excel file.

        :param fou: Path to the result file, including filename.
        :type fou: str
        :param sheet_name: Name of the sheet into the file, defaults to Excel default sheet name.
        :type sheet_name: str
        :param font_face: Font name used into the file, defaults to Excel default font face.
        :type font_face: str
        """
        if self._cursor:
            wb = openpyxl.Workbook()
            ws = wb.active
            if sheet_name:
                ws.title = sheet_name

            for col_num, cell in enumerate(self.row_header(), start=1):
                ws.cell(row=1, column=col_num).value = cell
                ws.cell(row=1, column=col_num).font = Font(name=font_face, bold=True)
                ws.cell(row=1, column=col_num).number_format = self._EXCEL_FORMATS[type(cell)]

            for row_num, row in enumerate(self._cursor, start=2):
                for col_num, cell in enumerate(row, start=1):
                    ws.cell(row=row_num, column=col_num).value = cell
                    ws.cell(row=row_num, column=col_num).font = Font(name=font_face)
                    ws.cell(row=row_num, column=col_num).number_format = self._EXCEL_FORMATS[type(cell)]

            ws.auto_filter.ref = ws.dimensions
            wb.save(fou)
